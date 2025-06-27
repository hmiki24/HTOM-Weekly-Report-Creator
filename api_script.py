from openai import AzureOpenAI
import gettoken
import openpyxl

#ご自身のapp-key記載↓
app_key = "xxxx" 
api_key = gettoken.get_access_token()


# Excelの読み込み
wb = openpyxl.load_workbook("weekly_SR.xlsx")


client = AzureOpenAI( 
    azure_endpoint = 'https://chat-ai.cisco.com', 
    api_key = api_key,
    api_version= "2024-12-01-preview" 
    ) 

#プロンプト
SUMMARIZE_PROMPT = """
You are an expert summarizer for support request (SR) for Nomura. 
Your task is to provide a concise and clear summary of the case, combining previous week's summary with recent updates.
The summary will be used for Nomura to briefly understand the flow of the SR.
Refer to customer as Nomura. Refer to Cisco TAC as TAC. Refer to any FE as FE.

Be concise on who did what, and the next action is pending from whom. 
Use specific phrases like:
- "TAC reviewed the logs and confirmed [parts] hardware failure."
- "FE successfully replaced [parts] which resolved the issue."
- "RMA dispatched for [parts] replacement, ETA set on [date]."
- "Pending Nomura for updates."
Always end with "Pending [who] to [what]."
When including date, format them as "day month" (e.g., "30 May" or "2 June").
Do not use indents.
"""

# Open Cases Sheet
ws_1 = wb["Current Open"]
for row in range(3, ws_1.max_row + 1): 
    title = ws_1.cell(row=row, column=2).value # ExcelのColumn B　読み込み
    last_week_summary = ws_1.cell(row=row, column=17).value  # ExcelのColumn Q　読み込み
    past_timeline = ws_1.cell(row=row, column=18).value # ExcelのColumn R　読み込み

    if past_timeline:
        print(f"Summarizing row {row}...") # ターミナルに進捗表示
        response = client.chat.completions.create( 
            model="gpt-4.1",
            messages = [
                {"role": "system", "content": SUMMARIZE_PROMPT},
                {"role": "user", "content": title},
                {"role": "user", "content": f"Last Week's Summary: {last_week_summary}"},
                {"role": "user", "content": f"Timeline: {past_timeline}"}
            ], 
            user=f'{{"appkey": "{app_key}"}}'
        )

        summary = response.choices[0].message.content.strip()
        existing_content = ws_1.cell(row=row, column=8).value  # 出力先のセルに値があれば、取得
        ws_1.cell(row=row, column=8).value = f"{existing_content}\n\n{summary}"  # ExcelのColumn Hに出力

# Save the updated file
wb.save('report_with_summary.xlsx')